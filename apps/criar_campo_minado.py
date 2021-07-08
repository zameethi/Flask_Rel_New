import os
import zipfile
from datetime import datetime
import openpyxl as xw
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import sqlite3
from vars_pandas import *
from shutil import copy
from apps.lista_ideal import listar_ideal


def zippar(arqs, origem):
    rel_zip = zipfile.ZipFile(f'{origem}/{arqs}.zip', 'w')
    for folder, subfolders, files in os.walk(f'{origem}'):
        for i, file in enumerate(files):
            if file.endswith('.xlsm') or file.endswith('.xlsx'):
                rel_zip.write(os.path.join(folder, file),
                              os.path.relpath(os.path.join(folder, file), f'{origem}'),
                              compress_type=zipfile.ZIP_DEFLATED)

    rel_zip.close()
    texto = rel_zip.filename
    return texto


def processar(id, concatenado, secoes, tabela, ordem, template):
    con = sqlite3.connect(os.path.abspath('apps/db/sortimento.db'))
    print(f'processando! secoes {secoes}')

    path = os.path.abspath(f'apps/db/Arquivos{id}')
    start = datetime.now()
	
    con.execute(f"UPDATE PROCESSO SET processando = 1 where id = {id};")
    con.commit()
    for root, dirs, files in os.walk(path):
        for name in files:
            if name.endswith('xlsm') or name.endswith('zip'):
                os.remove(os.path.join(root, name))

    for i, secao in enumerate(secoes):
        a = "'CAMPO MINADO CHECK'"
        b = "'CAMPO MINADO ALTERAÇÃO'"
        print(f'PR - {secao}')        
        if 'Telefonia' in template:
            df = pd.read_sql(f"""select * from sortemp where (codsortimento in (2021,2022,2023,2024,2025,2078,2201,2202,2203,2204,2205)) and CODSECAO = {secao} and DESCMERCADORIA not like '%CHIP PRE%' and DESCMERCADORIA not like '%CHIP POS%' and DESCMERCADORIA not like '%NANO CHIP%' and DESCMERCADORIA not like '%TRIPLO %' and DESCMERCADORIA not like '%SIM CARD %' and DESCMERCADORIA not like '%TIM CHIP %'
                                union 
                                select * from sortemp where (codsortimento >= 201 and codsortimento <=900)  and CODSECAO = {secao} and DESCMERCADORIA not like '%CHIP PRE%' and DESCMERCADORIA not like '%CHIP POS%' and DESCMERCADORIA not like '%NANO CHIP%' and DESCMERCADORIA not like '%TRIPLO %'and DESCMERCADORIA not like '%SIM CARD %' and DESCMERCADORIA not like '%TIM CHIP %';""", con)
        elif  secao == 15:            
            df = pd.read_sql(f"select * from sortemp where ((codsortimento in (2021,2022,2023,2024,2025,2078,2201,2202,2203,2204,2205)) or (codsortimento >= 201 and codsortimento <=900)) and CODSECAO = {secao};", con)
        
        else:
            df = pd.read_sql(f"select * from sortemp where ((codsortimento in (2021,2022,2023,2024,2025,2078,2201,2202,2203,2204,2205)) or (codsortimento >= 201 and codsortimento <=900)) and CODSECAO = {secao};", con)

        if df.empty:
            continue
        else:
            pass
        
        print('select realizado!')
        con.execute(f"UPDATE PROCESSO SET processando = 1 where id = {id};")
        con.commit()
        print('executado readsql!')

        progresso = (int(100.0 / float(len(secoes)) * float(i)))
        
        con.execute(f"UPDATE PROCESSO SET atual = {secao} where id = {id};")
        con.execute(f"UPDATE PROCESSO SET progresso = {progresso} where id = {id};")
        con.commit()


        file_name = datetime.strftime(datetime.today(), "%d-%m-%Y")

        new_cols = ['PERFIL', 'BANDEIRA', 'LOCAL_NOVO', 'LOCAL', 'TAMANHO', 'REGIÃO', 'REGIAO 2', 'CONCATENADO',
                    'CONCATENADO 2', 'CONCATENADO 3', 'CONCATENADO 4', 'CONCATENADO 5', 'CONCATENADO 6', 'CONCATENADO 7']
        for cols in new_cols:
            df[cols] = ''

        df['LOCAL_NOVO'] = df['DESCTAMANHO'].apply(lambda x: 'CONVENCIONAL' if 'RUA' in x else x)
        df['LOCAL_NOVO'] = df['LOCAL_NOVO'].apply(lambda x: 'CONVENCIONAL' if 'SHOPPING' in x else x)
        df['LOCAL_NOVO'] = df['LOCAL_NOVO'].apply(lambda x: 'QUIOSQUE' if 'QUIOSQUE' in x else x)
        df['LOCAL_NOVO'] = df['LOCAL_NOVO'].apply(lambda x: 'COMPACTA' if 'COMPACTA' in x else x)
        df['LOCAL_NOVO'] = df['LOCAL_NOVO'].apply(lambda x: 'DIGITAL' if 'DIGITAL' in x else x)

        df['LOCAL'] = df['DESCTAMANHO'].apply(lambda x: 'RUA' if 'RUA' in x else x)
        df['LOCAL'] = df['LOCAL'].apply(lambda x: 'SHOPPING' if 'SHOPPING' in x else x)
        df['LOCAL'] = df['LOCAL'].apply(lambda x: 'QUIOSQUE' if 'QUIOSQUE' in x else x)
        df['LOCAL'] = df['LOCAL'].apply(lambda x: 'COMPACTA' if 'COMPACTA' in x else x)
        df['LOCAL'] = df['LOCAL'].apply(lambda x: 'DIGITAL' if 'DIGITAL' in x else x)

        df['BANDEIRA'] = df['DESCTAMANHO'].apply(lambda x: 'CB' if '-CB-' in x else x)
        df['BANDEIRA'] = df['BANDEIRA'].apply(lambda x: 'PF' if '-PF-' in x else x)

        df['PERFIL'] = df['DESCTAMANHO'].apply(lambda x: 'PERFIL 1' if '-PERFIL 1-' in x else x)
        df['PERFIL'] = df['PERFIL'].apply(lambda x: 'PERFIL 2' if '-PERFIL 2-' in x else x)
        df['PERFIL'] = df['PERFIL'].apply(lambda x: 'PERFIL 3' if '-PERFIL 3-' in x else x)
        df['PERFIL'] = df['PERFIL'].apply(lambda x: 'PERFIL 4' if '-PERFIL 4-' in x else x)
        df['PERFIL'] = df['PERFIL'].apply(lambda x: 'PERFIL 5' if '-PERFIL 5-' in x else x)

        df['TAMANHO'] = df['DESCTAMANHO'].apply(lambda x: 'PP' if '-PP-' in x else x)
        df['TAMANHO'] = df['TAMANHO'].apply(lambda x: 'P' if '-P-' in x else x)
        df['TAMANHO'] = df['TAMANHO'].apply(lambda x: 'M' if '-M-' in x else x)
        df['TAMANHO'] = df['TAMANHO'].apply(lambda x: 'G' if '-G-' in x else x)
        df['TAMANHO'] = df['TAMANHO'].apply(lambda x: 'GG' if '-GG-' in x else x)
        df['TAMANHO'] = df['TAMANHO'].apply(lambda x: 'MEGA' if '-MEGA-' in x else x)

        df['REGIAO'] = df['DESCTAMANHO'].apply(lambda x: 'RJ/ES' if '-RJ/ES' in x else x)
        df['REGIAO'] = df['REGIAO'].apply(lambda x: 'SPI' if '-SPI' in x else x)
        df['REGIAO'] = df['REGIAO'].apply(lambda x: 'MG' if '-MG' in x else x)
        df['REGIAO'] = df['REGIAO'].apply(lambda x: 'SUL' if '-SUL' in x else x)
        df['REGIAO'] = df['REGIAO'].apply(lambda x: 'GDE SP' if '-GDE SP' in x else x)
        df['REGIAO'] = df['REGIAO'].apply(lambda x: 'CO/N' if '-CO/N' in x else x)
        df['REGIAO'] = df['REGIAO'].apply(lambda x: 'NE' if '-NE' in x else x)

        df['REGIAO 2'] = df['DESCTAMANHO'].apply(lambda x: 'RJ/ES' if '-RJ/ES' in x else x)
        df['REGIAO 2'] = df['REGIAO 2'].apply(lambda x: 'SPI' if '-SPI' in x else x)
        df['REGIAO 2'] = df['REGIAO 2'].apply(lambda x: 'MG' if '-MG' in x else x)
        df['REGIAO 2'] = df['REGIAO 2'].apply(lambda x: 'SUL' if '-SUL' in x else x)
        df['REGIAO 2'] = df['REGIAO 2'].apply(lambda x: 'GDE SP' if '-GDE SP' in x else x)
        df['REGIAO 2'] = df['REGIAO 2'].apply(lambda x: 'CO/N' if '-CO/N' in x else x)
        df['REGIAO 2'] = df['REGIAO 2'].apply(lambda x: 'NE' if '-NE' in x else x)

        df['CONCATENADO'] = df['LOCAL_NOVO'].map(str) + '-' + df['PERFIL'].map(str) + '-' + df['TAMANHO'].map(str)
        df['CONCATENADO 2'] = df['LOCAL_NOVO'].map(str) + '-' + df['PERFIL'].map(str) + '-' + df['REGIAO'].map(str) + '-' + df['TAMANHO'].map(str)
        df['CONCATENADO 3'] = df['LOCAL_NOVO'].map(str) + '-' + df['PERFIL'].map(str) + '-' + df['REGIAO'].map(str)
        df['CONCATENADO 4'] = df['LOCAL'].map(str) + '-' + df['TAMANHO'].map(str)
        df['CONCATENADO 5'] = df['LOCAL_NOVO'].map(str) + '-' + df['PERFIL'].map(str) + '-' + df['BANDEIRA'].map(str) + '-' + df['TAMANHO'].map(str)
        df['CONCATENADO 6'] = df['PERFIL'].map(str) + '-' + df['BANDEIRA'].map(str) + '-' + df['LOCAL'] + '-' + df['REGIAO'].map(str)
        df['CONCATENADO 7'] = df['PERFIL'].map(str) + '-' + df['BANDEIRA'].map(str) + '-' + df['LOCAL'] + '-' + df['REGIAO'].map(str)

        df_all = df.copy()
        sortemp = df.copy()
        
        df = df.sort_values('CODMERCADORIA', ascending=True)

        idx = df['CODMERCADORIA'].drop_duplicates().index
        df = df.loc[idx, :]

        df1 = pd.DataFrame(columns=tabela)

        df1['alteracao'] = ''
        df1['SKU'] = df['CODMERCADORIA'].copy()
        df1['Descrição'] = df['DESCMERCADORIA'].copy()
        df1['Voltagem'] = df['VOLTAGEM'].copy()
        df1['Fornecedor'] = df['DESCMARCA']
        df1['Situação'] = df['STATUS'].copy()
        df1['Espécie'] = df['DESCESPECIE'].copy()
        df1['check'] = ''
        df1['obg'] = ''
        df1['sug'] = ''
        df1['total'] = ''
        df1 = df1.reset_index(drop=True)

        n_rows = 16

        # iterating over indices
        for col in (df1.index):
            n_rows += 1
            df1.loc[col, 'obg'] = f'=SUMIF(L{n_rows}:GN{n_rows},1,$L$14:$GN$14)'
            df1.loc[
                col, 'check'] = f'=CONCATENATE(L{n_rows},M{n_rows},N{n_rows},O{n_rows},P{n_rows},Q{n_rows},R{n_rows},S{n_rows},T{n_rows},U{n_rows},V{n_rows},W{n_rows},X{n_rows},Y{n_rows},Z{n_rows},AA{n_rows},AB{n_rows},AC{n_rows},AD{n_rows},AE{n_rows},AF{n_rows},AG{n_rows},AH{n_rows},AI{n_rows},AJ{n_rows},AK{n_rows},AL{n_rows},AM{n_rows},AN{n_rows},AO{n_rows},AP{n_rows},AQ{n_rows},AR{n_rows},AS{n_rows},AT{n_rows},AU{n_rows},AV{n_rows},AW{n_rows},AX{n_rows},AY{n_rows},AZ{n_rows},BA{n_rows},BB{n_rows},BC{n_rows},BD{n_rows},BE{n_rows},BF{n_rows},BG{n_rows},BH{n_rows},BI{n_rows},BJ{n_rows},BK{n_rows},BL{n_rows},BM{n_rows},BN{n_rows},BO{n_rows},BP{n_rows},BQ{n_rows},BR{n_rows},BS{n_rows},BT{n_rows},BU{n_rows},BV{n_rows},BW{n_rows},BX{n_rows},BY{n_rows},BZ{n_rows},CA{n_rows},CB{n_rows},CC{n_rows},CD{n_rows},CE{n_rows},CF{n_rows},CG{n_rows},CH{n_rows},CI{n_rows},CJ{n_rows},CK{n_rows},CL{n_rows},CM{n_rows},CN{n_rows},CO{n_rows},CP{n_rows},CQ{n_rows},CR{n_rows},CS{n_rows},CT{n_rows},CU{n_rows},CV{n_rows},CW{n_rows},CX{n_rows},CY{n_rows},CZ{n_rows},DA{n_rows},DB{n_rows},DC{n_rows},DD{n_rows},DE{n_rows},DF{n_rows},DG{n_rows},DH{n_rows},DI{n_rows},DJ{n_rows},DK{n_rows},DL{n_rows},DM{n_rows},DN{n_rows},DO{n_rows},DP{n_rows},DQ{n_rows},DR{n_rows},DS{n_rows},DT{n_rows},DU{n_rows},DV{n_rows},DW{n_rows},DX{n_rows},DY{n_rows},DZ{n_rows},EA{n_rows},EB{n_rows},EC{n_rows},ED{n_rows},EE{n_rows},EF{n_rows},EG{n_rows},EH{n_rows},EI{n_rows},EJ{n_rows},EK{n_rows},EL{n_rows},EM{n_rows},EN{n_rows},EO{n_rows},EP{n_rows},EQ{n_rows},ER{n_rows},ES{n_rows},ET{n_rows},EU{n_rows},EV{n_rows},EW{n_rows},EX{n_rows},EY{n_rows},EZ{n_rows},FA{n_rows},FB{n_rows},FC{n_rows},FD{n_rows},FE{n_rows},FF{n_rows},FG{n_rows},FH{n_rows},FI{n_rows},FJ{n_rows},FK{n_rows},FL{n_rows},FM{n_rows},FN{n_rows},FO{n_rows},FP{n_rows},FQ{n_rows},FR{n_rows},FS{n_rows},FT{n_rows},FU{n_rows},FV{n_rows},FW{n_rows},FX{n_rows},FY{n_rows},FZ{n_rows},GA{n_rows},GB{n_rows},GC{n_rows},GD{n_rows},GE{n_rows},GF{n_rows},GG{n_rows},GH{n_rows},GI{n_rows},GJ{n_rows},GK{n_rows},GL{n_rows},GM{n_rows},GN{n_rows})'
            df1.loc[
                col, 'alteracao'] = f'=IF(AND(B{n_rows}="",IFERROR(IF(VLOOKUP(B{n_rows},{a}!B:H,7,0)=H{n_rows},"Não","Sim"),1)=1),"",IFERROR(IF(VLOOKUP(B{n_rows},{a}!B:H,7,0)=H{n_rows},"Não","Sim"),"Novo Produto"))'
            df1.loc[col, 'sug'] = f'=SUMIF(L{n_rows}:GN{n_rows},3,$L$14:$GN$14)'
            df1.loc[col, 'total'] = f'=I{n_rows}+J{n_rows}'


        df1 = df1.copy()
        colorder = ['CODSORTIMENTO', 'CODSECAO', 'DESCSECAO', 'CODTAMANHO', 'DESCTAMANHO',
                    'CODPUBLICO', 'DESCPUBLICO', 'CODMERCADORIA', 'DESCMERCADORIA', 'MERCCONJ',
                    'PRIORIDADE', 'QTDE', 'VOLTAGEM', 'VIGENCIAINICIO', 'VIGENCIAFIM', 'PERFIL',
                    'BANDEIRA', 'LOCAL', 'LOCAL_NOVO', 'TAMANHO', 'REGIÃO', 'STATUS', 'DESCMARCA',
                    'DESCESPECIE', 'CONCATENADO', 'CONCATENADO 2']
        df = df[colorder]
        sortemp = sortemp[colorder]


        df1 = df1[ordem]
        df_all = df_all[['CODMERCADORIA', 'PRIORIDADE', f'{concatenado}']].sort_values('CODMERCADORIA')
        df1.reset_index(drop=True, inplace=True)
        print('começando do looping')
        for n, data in enumerate(df1.itertuples()):
            for item in df_all.itertuples():
                if data[2] == item[1]:
                    for x, i in enumerate(tabela):
                        if item[3] == i:
                            try:
                                df1.loc[data.Index, f"{i}"] = int(item[2])
                            except:
                                df1.iloc[data.Index, x] = int(item[2])
            
        print('montando df!')
        df1.loc[df1.index.max() + 1] = None
        df1.iloc[df1.index.max(), 0] = f'=IF(AND(B{n_rows}="",IFERROR(IF(VLOOKUP(B{n_rows},{a}!B:H,7,0)=H{n_rows},"Não","Sim"),1)=1),"",IFERROR(IF(VLOOKUP(B{n_rows},{a}!B:H,7,0)=H{n_rows},"Não","Sim"),"Novo Produto"))'
        df1.iloc[df1.index.max(), 1] = 'Total Geral'


        sortim = pd.read_sql(f"select * from sortim where ((codsortimento in (2021,2022,2023,2024,2025,2078)) or (codsortimento >= 201 and codsortimento <=900)) and CODSECAO = {secao};", con)
        print('Sql sortim!')

        new_cols = ['PERFIL', 'BANDEIRA', 'LOCAL_NOVO', 'LOCAL', 'TAMANHO', 'REGIÃO', 'REGIAO 2', 'CONCATENADO',
                    'CONCATENADO 2', 'CONCATENADO 3', 'CONCATENADO 4', 'CONCATENADO 5', 'CONCATENADO 6', 'CONCATENADO 7']
        for i in new_cols:
            sortim[i] = ''

        sortim['LOCAL_NOVO'] = sortim['DESCTAMANHO'].apply(lambda x: 'CONVENCIONAL' if 'RUA' in x else x)
        sortim['LOCAL_NOVO'] = sortim['LOCAL_NOVO'].apply(lambda x: 'CONVENCIONAL' if 'SHOPPING' in x else x)
        sortim['LOCAL_NOVO'] = sortim['LOCAL_NOVO'].apply(lambda x: 'QUIOSQUE' if 'QUIOSQUE' in x else x)
        sortim['LOCAL_NOVO'] = sortim['LOCAL_NOVO'].apply(lambda x: 'COMPACTA' if 'COMPACTA' in x else x)
        sortim['LOCAL_NOVO'] = sortim['LOCAL_NOVO'].apply(lambda x: 'DIGITAL' if 'DIGITAL' in x else x)

        sortim['LOCAL'] = sortim['DESCTAMANHO'].apply(lambda x: 'RUA' if 'RUA' in x else x)
        sortim['LOCAL'] = sortim['LOCAL'].apply(lambda x: 'SHOPPING' if 'SHOPPING' in x else x)
        sortim['LOCAL'] = sortim['LOCAL'].apply(lambda x: 'QUIOSQUE' if 'QUIOSQUE' in x else x)
        sortim['LOCAL'] = sortim['LOCAL'].apply(lambda x: 'COMPACTA' if 'COMPACTA' in x else x)
        sortim['LOCAL'] = sortim['LOCAL'].apply(lambda x: 'DIGITAL' if 'DIGITAL' in x else x)

        sortim['BANDEIRA'] = sortim['DESCTAMANHO'].apply(lambda x: 'CB' if '-CB-' in x else x)
        sortim['BANDEIRA'] = sortim['BANDEIRA'].apply(lambda x: 'PF' if '-PF-' in x else x)

        sortim['PERFIL'] = sortim['DESCTAMANHO'].apply(lambda x: 'PERFIL 1' if '-PERFIL 1-' in x else x)
        sortim['PERFIL'] = sortim['PERFIL'].apply(lambda x: 'PERFIL 2' if '-PERFIL 2-' in x else x)
        sortim['PERFIL'] = sortim['PERFIL'].apply(lambda x: 'PERFIL 3' if '-PERFIL 3-' in x else x)
        sortim['PERFIL'] = sortim['PERFIL'].apply(lambda x: 'PERFIL 4' if '-PERFIL 4-' in x else x)
        sortim['PERFIL'] = sortim['PERFIL'].apply(lambda x: 'PERFIL 5' if '-PERFIL 5-' in x else x)

        sortim['TAMANHO'] = sortim['DESCTAMANHO'].apply(lambda x: 'PP' if '-PP-' in x else x)
        sortim['TAMANHO'] = sortim['TAMANHO'].apply(lambda x: 'P' if '-P-' in x else x)
        sortim['TAMANHO'] = sortim['TAMANHO'].apply(lambda x: 'M' if '-M-' in x else x)
        sortim['TAMANHO'] = sortim['TAMANHO'].apply(lambda x: 'G' if '-G-' in x else x)
        sortim['TAMANHO'] = sortim['TAMANHO'].apply(lambda x: 'GG' if '-GG-' in x else x)
        sortim['TAMANHO'] = sortim['TAMANHO'].apply(lambda x: 'MEGA' if '-MEGA-' in x else x)

        sortim['REGIAO'] = sortim['DESCTAMANHO'].apply(lambda x: 'RJ/ES' if '-RJ/ES' in x else x)
        sortim['REGIAO'] = sortim['REGIAO'].apply(lambda x: 'SPI' if '-SPI' in x else x)
        sortim['REGIAO'] = sortim['REGIAO'].apply(lambda x: 'MG' if '-MG' in x else x)
        sortim['REGIAO'] = sortim['REGIAO'].apply(lambda x: 'SUL' if '-SUL' in x else x)
        sortim['REGIAO'] = sortim['REGIAO'].apply(lambda x: 'GDE SP' if '-GDE SP' in x else x)
        sortim['REGIAO'] = sortim['REGIAO'].apply(lambda x: 'CO/N' if '-CO/N' in x else x)
        sortim['REGIAO'] = sortim['REGIAO'].apply(lambda x: 'NE' if '-NE' in x else x)

        sortim['REGIAO 2'] = sortim['DESCTAMANHO'].apply(lambda x: 'RJ/ES' if '-RJ/ES' in x else x)
        sortim['REGIAO 2'] = sortim['REGIAO 2'].apply(lambda x: 'SPI' if '-SPI' in x else x)
        sortim['REGIAO 2'] = sortim['REGIAO 2'].apply(lambda x: 'MG' if '-MG' in x else x)
        sortim['REGIAO 2'] = sortim['REGIAO 2'].apply(lambda x: 'SUL' if '-SUL' in x else x)
        sortim['REGIAO 2'] = sortim['REGIAO 2'].apply(lambda x: 'GDE SP' if '-GDE SP' in x else x)
        sortim['REGIAO 2'] = sortim['REGIAO 2'].apply(lambda x: 'CO/N' if '-CO/N' in x else x)
        sortim['REGIAO 2'] = sortim['REGIAO 2'].apply(lambda x: 'NE' if '-NE' in x else x)

        sortim['CONCATENADO'] = sortim['LOCAL_NOVO'].map(str) + '-' + sortim['PERFIL'].map(str) + '-' + sortim[
            'TAMANHO'].map(str)
        sortim['CONCATENADO 2'] = sortim['LOCAL_NOVO'].map(str) + '-' + sortim['PERFIL'].map(str) + '-' + sortim[
            'REGIAO'].map(str) + '-' + sortim['TAMANHO'].map(str)
        sortim['CONCATENADO 3'] = sortim['LOCAL_NOVO'].map(str) + '-' + sortim['PERFIL'].map(str) + '-' + sortim[
            'REGIAO'].map(str)
        sortim['CONCATENADO 4'] = sortim['LOCAL'].map(str) + '-' + sortim['TAMANHO'].map(str)
        sortim['CONCATENADO 5'] = sortim['LOCAL_NOVO'].map(str) + '-' + sortim['PERFIL'].map(str) + '-' + sortim[
            'BANDEIRA'].map(str) + '-' + sortim['TAMANHO'].map(str)
        sortim['CONCATENADO 6'] = sortim['PERFIL'].map(str) + '-' + sortim['BANDEIRA'].map(str) + '-' + sortim[
            'LOCAL'] + '-' + sortim['REGIAO'].map(str)
        sortim['CONCATENADO 7'] = sortim['PERFIL'].map(str) + '-' + sortim['BANDEIRA'].map(str) + '-' + sortim[
            'LOCAL'] + '-' + sortim['REGIAO'].map(str)
        colorder = ['CODPUBLICO', 'DESCPUBLICO', 'CODSECAO', 'DESCSECAO', 'CODSORTIMENTO', 'CODTAMANHO', 'DESCTAMANHO',
                    'PERFIL', 'BANDEIRA', 'LOCAL_NOVO', 'LOCAL', 'TAMANHO', 'REGIÃO', 'REGIAO 2', 'CONCATENADO',
                    'CONCATENADO 2', 'CONCATENADO 3', 'CONCATENADO 4', 'CONCATENADO 5', 'CONCATENADO 6', 'CONCATENADO 7', 'EMPRESA',
                    'FILIAL', 'TIPOPUBLICO', 'SOLICVOLTDIF']
        sortim = sortim[colorder]
    
        n_ideal = listar_ideal()
        

        print('iniciando excel!')
        xl = xw.load_workbook(os.path.abspath(f'apps/db/{template}.xlsm'), keep_vba=True)

        sheets = xl.sheetnames

        w1 = xl[sheets[0]]
        w2 = xl[sheets[1]]
        w3 = xl[sheets[2]]
        w4 = xl[sheets[4]]
       
        df1.iloc[df1.index.max(), 0] = None
        df1.iloc[df1.index.max(), 1] = None

        print('dataframete to rows df1!')
        df1 = dataframe_to_rows(df1, index=False, header=False)
        for n_row, linhas in enumerate(df1, start=17):
            for n_cols,cols in enumerate(linhas, start=1):
                w1.cell(row=n_row, column=n_cols).value = cols
                w2.cell(row=n_row, column=n_cols).value = cols

        print('dataframete to rows sortemp!')
        sortemp = dataframe_to_rows(sortemp, index=False, header=False)
        for n_row, linhas in enumerate(sortemp, start=2):
            for n_cols,cols in enumerate(linhas, start=1):
                w3.cell(row=n_row, column=n_cols).value = cols
        
        print('dataframete to rows sortim!')
        sortim = dataframe_to_rows(sortim, index=False, header=False)
        for n_row, linhas in enumerate(sortim, start=2):
            for n_cols,cols in enumerate(linhas, start=1):
                w4.cell(row=n_row, column=n_cols).value = cols

        if secao in n_ideal.keys():
            if template in ('Perfil_Tamanho','Perfil_Regiao'):
                ide = 7
            else:
                ide = 8
 
            for cols, valor in enumerate(n_ideal[secao], start=12):
                w1.cell(row=ide, column=n_cols).value = valor
                w2.cell(row=ide, column=n_cols).value = valor
        

        print('salvando!')
        xl.save(f'{path}/secao-{secao}_{file_name}.xlsm')
        xl.close()
        

        del df, df_all, df1, sortim, sortemp
        gc.collect()
    
    
    con.execute(f"UPDATE PROCESSO SET processando = 2 where id = {id};")
    con.execute(f"UPDATE PROCESSO SET atual = 0 where id = {id};")
    con.execute(f"UPDATE PROCESSO SET progresso = 100 where id = {id};")
    con.commit()
        
    f = f'{template}-{datetime.strftime(datetime.today(), "%d-%m-%Y")}'
    
    copy(os.path.abspath('apps/db/DE PARA MACRO.xlsx'), os.path.abspath(f'apps/db/Arquivos{id}'))
    
    
    zippar(f, os.path.abspath(f'apps/db/Arquivos{id}'))

    stop = datetime.now() - start
    con.execute(f"UPDATE PROCESSO SET tempo = '{str(stop)[:-7]} - finalizado: {datetime.strftime(datetime.now(), '%d/%m/%Y %H:%M')}' where id = {id};")
    con.commit()
    con.close()
    print(f'Time:, {stop}')

    return True
